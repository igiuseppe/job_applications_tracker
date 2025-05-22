import requests
from bs4 import BeautifulSoup
import time
import json
import csv
import os
import datetime
import re
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import config

def get_job_description(job_public_url):
    """Fetch a job description from a public LinkedIn URL"""
    response = requests.get(job_public_url)
    soup = BeautifulSoup(response.text, "html.parser")
    description = soup.find('div', class_='description__text description__text--rich').text.strip()  # Example: Job title
    return description

# Helper function to safely get text
def get_text_or_none(element):
    return element.get_text(strip=True) if element else None

def get_href_or_none(element):
    return element['href'] if element and element.has_attr('href') else None

def clean_job_html(html_content, work_type=None):
    """Extract structured job data from LinkedIn job HTML"""
    soup = BeautifulSoup(html_content, 'lxml')

    # --- Top Card Information ---
    job_title_element = soup.find('h2', class_='top-card-layout__title')
    job_title = get_text_or_none(job_title_element)

    company_name_element = soup.find('a', class_='topcard__org-name-link')
    company_name = get_text_or_none(company_name_element)

    location_element = soup.find('span', class_='topcard__flavor--bullet')
    location = get_text_or_none(location_element) # This might need refinement if there are multiple such spans

    posted_time_ago_element = soup.find('span', class_='posted-time-ago__text')
    posted_time_ago = get_text_or_none(posted_time_ago_element)

    num_applicants_note_element = soup.find('figcaption', class_='num-applicants__caption')
    num_applicants_note = get_text_or_none(num_applicants_note_element)

    see_who_hired_text_element = soup.find('p', class_='face-pile__text')
    see_who_hired_text = get_text_or_none(see_who_hired_text_element)

    # --- Recruiter Information ---
    recruiter_message_element = soup.select_one('.message-the-recruiter > p')
    recruiter_message = get_text_or_none(recruiter_message_element)

    recruiter_name_element = soup.find('h3', class_='base-main-card__title--link')
    recruiter_name = get_text_or_none(recruiter_name_element)

    recruiter_tagline_element = soup.find('h4', class_='base-main-card__subtitle')
    recruiter_tagline = get_text_or_none(recruiter_tagline_element)


    # --- Job Description ---
    job_description_container = soup.find('div', class_='show-more-less-html__markup')
    if job_description_container:
        job_description_paragraphs = job_description_container.find_all(['p', 'ul'])
        job_description_parts = []
        for item in job_description_paragraphs:
            if item.name == 'ul':
                list_items = [li.get_text(separator=' ', strip=True) for li in item.find_all('li')]
                job_description_parts.append("\n".join(["- " + li for li in list_items]))
            else:
                job_description_parts.append(item.get_text(separator=' ', strip=True))
        job_description = "\n\n".join(job_description_parts)
    else:
        job_description = None

    # --- Job Criteria ---
    seniority_level = None
    employment_type = None
    job_function = None
    industries = None

    criteria_items = soup.find_all('li', class_='description__job-criteria-item')
    for item in criteria_items:
        header = item.find('h3', class_='description__job-criteria-subheader')
        text_element = item.find('span', class_='description__job-criteria-text--criteria')
        if header and text_element:
            header_text = get_text_or_none(header)
            criteria_text = get_text_or_none(text_element)
            if header_text == 'Seniority level':
                seniority_level = criteria_text
            elif header_text == 'Employment type':
                employment_type = criteria_text
            elif header_text == 'Job function':
                job_function = criteria_text
            elif header_text == 'Industries':
                industries = criteria_text

    # --- Referral Information ---
    referral_text_element = soup.select_one('.find-a-referral__cta-container > p')
    referral_text = get_text_or_none(referral_text_element)

    referral_cta_element = soup.select_one('a.find-a-referral__cta')
    referral_cta_text = get_text_or_none(referral_cta_element)

    job_title_anchor_element = soup.find('a', class_='topcard__link', attrs={'data-tracking-control-name': 'public_jobs_topcard-title'})
    job_link = get_href_or_none(job_title_anchor_element)

    company_name_anchor_element = soup.find('a', class_='topcard__org-name-link', attrs={'data-tracking-control-name': 'public_jobs_topcard-org-name'})
    company_link = get_href_or_none(company_name_anchor_element)

    recruiter_profile_anchor_element = soup.select_one('div.message-the-recruiter a.base-card__full-link')
    # This selector targets the main link covering the recruiter's card.
    # An alternative, more specific to the recruiter's name if the structure is consistent:
    # recruiter_profile_anchor_element = soup.select_one('h3.base-main-card__title--link a') # if the <a> was inside the <h3>
    # However, in your HTML, the <a> wraps the content.
    recruiter_link = get_href_or_none(recruiter_profile_anchor_element)

    # Extract the publishing date from posted_time_ago
    publishing_date = extract_publishing_date(posted_time_ago)
    
    # Get work type name if specified
    work_type_name = config.WORK_TYPE_NAMES.get(work_type, "Unknown")

    # Create and return a dictionary with all extracted variables
    job_data = {
        "job_title": job_title,
        "company_name": company_name,
        "location": location,
        "posted_time_ago": posted_time_ago,
        "publishing_date": publishing_date,
        "num_applicants_note": num_applicants_note,
        "recruiter_message": recruiter_message,
        "recruiter_name": recruiter_name,
        "recruiter_tagline": recruiter_tagline,
        "job_description": job_description,
        "seniority_level": seniority_level,
        "employment_type": employment_type,
        "job_function": job_function,
        "industries": industries,
        "job_link": job_link,
        "company_link": company_link,
        "recruiter_link": recruiter_link,
        "work_type": work_type,
        "work_type_name": work_type_name
    }
    
    return job_data

def extract_publishing_date(posted_time_ago):
    """
    Extract an approximate publishing date from the 'posted X time ago' text
    """
    if not posted_time_ago:
        return datetime.datetime.now().strftime("%Y-%m-%d")
    
    today = datetime.datetime.now()
    
    # Handle different time formats
    if "minute" in posted_time_ago or "hour" in posted_time_ago:
        return today.strftime("%Y-%m-%d")
    elif "day" in posted_time_ago:
        match = re.search(r'(\d+)', posted_time_ago)
        if match:
            days = int(match.group(1))
            date = today - datetime.timedelta(days=days)
            return date.strftime("%Y-%m-%d")
    elif "week" in posted_time_ago:
        match = re.search(r'(\d+)', posted_time_ago)
        if match:
            weeks = int(match.group(1))
            date = today - datetime.timedelta(weeks=weeks)
            return date.strftime("%Y-%m-%d")
    elif "month" in posted_time_ago:
        match = re.search(r'(\d+)', posted_time_ago)
        if match:
            months = int(match.group(1))
            # Approximate months as 30 days
            date = today - datetime.timedelta(days=months*30)
            return date.strftime("%Y-%m-%d")
    
    # Default to today if we can't parse
    return today.strftime("%Y-%m-%d")

def get_job_list_page(keywords, location, geoId, start_position, work_type):
    """
    Fetches a page of job listings from LinkedIn
    
    Args:
        keywords: Search keywords
        location: Location to search in
        geoId: LinkedIn GeoID for the location
        start_position: Starting position for pagination
        work_type: Work type filter (None, 1=on-site, 2=remote, 3=hybrid)
    """

    
    list_url = config.LINKEDIN_JOB_LIST_URL_TEMPLATE.format(
        keywords=keywords,
        location=location,
        geoId=geoId,
        work_type=work_type,
        start_position=start_position
    )
    
    print(f"Fetching URL: {list_url}")  # Debugging output
    
    try:
        response = requests.get(list_url)
        soup = BeautifulSoup(response.text, "html.parser")
        page_jobs = soup.find_all("li")
        return page_jobs
    except Exception as e:
        print(f"Error fetching job list page: {str(e)}")
        return []

def extract_job_id(job_element):
    """Extracts job ID from a job listing element"""
    base_card_div = job_element.find("div", {"class": "base-card"})
    if not base_card_div or not base_card_div.get("data-entity-urn"):
        return None
    
    return base_card_div.get("data-entity-urn").split(":")[-1]

def fetch_job_details(job_id, work_type=None):
    """Fetches and processes details for a specific job"""
    job_url = config.LINKEDIN_JOB_DETAIL_URL_TEMPLATE.format(job_id=job_id)
    
    try:
        job_response = requests.get(job_url)
        html_content = job_response.text
        job_details = clean_job_html(html_content, work_type)
        
        # Add job_id to the details
        job_details["job_id"] = job_id
        return job_details
    except Exception as e:
        print(f"Error scraping job {job_id}: {str(e)}")
        return None

def save_jobs_to_file(jobs, filename="output/linkedin_jobs.json"):
    """Saves job data to a JSON file"""
    # Ensure output directory exists
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(jobs, f, ensure_ascii=False, indent=2)
    print(f"Job data saved to {filename}")

def load_existing_jobs_from_crm(excel_path="output/jobs_crm.xlsx"):
    """
    Load existing jobs from the CRM Excel file
    Returns a set of job IDs and a list of existing job records
    """
    existing_job_ids = set()
    existing_jobs = []
    
    if not os.path.exists(excel_path):
        return existing_job_ids, existing_jobs
    
    try:
        # Read Excel file
        df = pd.read_excel(excel_path)
        
        # Convert DataFrame to list of dictionaries
        existing_jobs = df.to_dict('records')
        
        # Create set of job IDs
        for job in existing_jobs:
            if 'job_id' in job and job['job_id']:
                existing_job_ids.add(str(job['job_id']))
                
        print(f"Loaded {len(existing_jobs)} existing jobs from CRM")
        return existing_job_ids, existing_jobs
    
    except Exception as e:
        print(f"Error loading existing jobs from Excel: {str(e)}")
        return set(), []

def update_jobs_crm(new_jobs, excel_path="output/jobs_crm.xlsx"):
    """
    Update the CRM Excel file with new job listings
    - Adds only new jobs (not already in CRM)
    - Orders jobs by publishing date (newest first)
    """
    # Ensure output directory exists
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    # Get existing jobs
    existing_job_ids, existing_jobs = load_existing_jobs_from_crm(excel_path)
    
    # Add only new jobs
    added_count = 0
    for job in new_jobs:
        if job['job_id'] not in existing_job_ids:
            # Add status fields for CRM functionality
            job['status'] = 'New'
            job['notes'] = ''
            job['date_added'] = datetime.datetime.now().strftime("%Y-%m-%d")
            existing_jobs.append(job)
            added_count += 1
    
    if added_count == 0:
        print("No new jobs to add to CRM")
        return
    
    # Sort all jobs by publishing date (newest first)
    sorted_jobs = sorted(existing_jobs, 
                         key=lambda x: x.get('publishing_date', '1970-01-01'), 
                         reverse=True)
    
    # Define fields for the Excel file
    fieldnames = [
        'job_id', 'job_title', 'company_name', 'location', 'publishing_date',
        'posted_time_ago', 'seniority_level', 'employment_type', 'job_function',
        'industries', 'status', 'notes', 'date_added', 'job_link'
    ]
    
    # Create a DataFrame with only the fields we want to include
    rows = []
    for job in sorted_jobs:
        # Only include fields that are in fieldnames
        row = {k: job.get(k, '') for k in fieldnames}
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    # Write to Excel
    try:
        # Create a styled Excel writer
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Job Listings')
            
            # Access the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['Job Listings']
            
            # Format headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column width
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max(max_length + 2, 10)
                worksheet.column_dimensions[column].width = min(adjusted_width, 50)
            
            # Add filters to headers
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Color coding for status
            for row_idx, row in enumerate(df.iterrows(), 2):  # Start from row 2 (after header)
                status = row[1].get('status', '')
                status_cell = worksheet.cell(row=row_idx, column=fieldnames.index('status') + 1)
                
                if status == 'New':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif status == 'Applied':
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif status == 'Interview':
                    status_cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
                elif status == 'Rejected':
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif status == 'Offer':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        print(f"Added {added_count} new jobs to CRM. Total jobs in CRM: {len(sorted_jobs)}")
        print(f"CRM data saved to {excel_path}")
        
    except Exception as e:
        print(f"Error saving to Excel file: {str(e)}")
        # Fallback to CSV if Excel save fails
        csv_path = excel_path.replace('.xlsx', '.csv')
        pd.DataFrame(rows).to_csv(csv_path, index=False)
        print(f"Saved to CSV instead at {csv_path}")

def scrape_linkedin_jobs(keywords, location, geoId, work_type, jobs_per_page, max_pages):
    """
    Main function to scrape LinkedIn jobs with pagination
    
    Args:
        keywords: Search keywords
        location: Location to search in
        geoId: LinkedIn GeoID for the location
        work_type: Work type filter (None, 1=on-site, 2=remote, 3=hybrid)
        jobs_per_page: Number of jobs per page
        max_pages: Maximum number of pages to scrape
    """
    all_jobs = []
    page = 0
    
    # Get work type name for display
    work_type_name = config.WORK_TYPE_NAMES.get(work_type, "Any")
    print(f"Searching for {work_type_name} jobs: {keywords} in {location} (work_type value: {work_type})")
    
    while page < max_pages:
        start_position = page * jobs_per_page
        print(f"Scraping page {page+1}, jobs {start_position+1}-{start_position+jobs_per_page}")
        
        page_jobs = get_job_list_page(keywords, location, geoId, start_position, work_type)
        
        # If no jobs are found, we've reached the end
        if not page_jobs:
            print(f"No more jobs found after {len(all_jobs)} jobs")
            break
            
        print(f"Found {len(page_jobs)} jobs on page {page+1}")
        
        for job in page_jobs:
            job_id = extract_job_id(job)
            if not job_id:
                continue
            
            job_details = fetch_job_details(job_id, work_type)
            if job_details:
                # Add search parameters to the job details
                job_details["search_keywords"] = keywords
                job_details["search_location"] = location
                job_details["search_geoId"] = geoId
                
                # Ensure work_type and work_type_name are properly set
                job_details["work_type"] = work_type
                job_details["work_type_name"] = work_type_name
                
                all_jobs.append(job_details)
                print(f"Scraped: {job_details['job_title']} at {job_details['company_name']}")
            
            # Add a delay to avoid rate limiting
            time.sleep(config.DELAY_BETWEEN_REQUESTS)
        
        page += 1
    
    print(f"Total jobs scraped: {len(all_jobs)}")
    return all_jobs

def main():
    # Search parameters
    keywords = "Business%2BAnalyst"
    location = "Italia"
    geoId = "103350119"
    max_pages = 1
    
    # Scrape jobs
    all_jobs = scrape_linkedin_jobs(keywords, location, geoId, max_pages=max_pages)
    
    # Save results to JSON
    save_jobs_to_file(all_jobs)
    
    # Update CRM with new jobs
    update_jobs_crm(all_jobs)

if __name__ == "__main__":
    main()
