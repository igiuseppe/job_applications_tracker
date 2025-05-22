import os
import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import config

def load_existing_jobs_from_crm(excel_path=config.CRM_EXCEL_PATH):
    """
    Load existing jobs from the CRM Excel file
    Returns a set of job IDs and a list of existing job records
    """
    existing_job_ids = set()
    existing_jobs = []
    
    if not os.path.exists(excel_path):
        return existing_job_ids, existing_jobs
    
    try:
        # Read Excel file - we need to read all sheets except 'Summary'
        xl = pd.ExcelFile(excel_path)
        sheet_names = [s for s in xl.sheet_names if s != 'Summary']
        
        # Read each sheet
        for sheet_name in sheet_names:
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            
            # Convert DataFrame to list of dictionaries
            sheet_jobs = df.to_dict('records')
            
            # Add to existing jobs list
            existing_jobs.extend(sheet_jobs)
        
        # Create set of job IDs
        for job in existing_jobs:
            if 'job_id' in job and job['job_id'] and not pd.isna(job['job_id']):
                existing_job_ids.add(str(job['job_id']))
                
        print(f"Loaded {len(existing_jobs)} existing jobs from CRM across {len(sheet_names)} sheets")
        return existing_job_ids, existing_jobs
    
    except Exception as e:
        print(f"Error loading existing jobs from Excel: {str(e)}")
        return set(), []

def update_jobs_crm(new_jobs, excel_path=config.CRM_EXCEL_PATH):
    """
    Update the CRM Excel file with new job listings
    - Adds only new jobs (not already in the same search sheet)
    - Orders jobs by publishing date (newest first)
    - Creates separate sheets for different search combinations
    """
    # Ensure output directory exists
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    # Group the new jobs by search parameters
    new_job_groups = {}
    
    for job in new_jobs:
        # Create a key based on search parameters
        keywords = job.get('search_keywords', 'Unknown')
        location = job.get('search_location', 'Unknown')
        work_type = job.get('work_type', None)
        
        # Simplified key and sheet name
        group_key = f"{keywords}_{location}_{work_type}"
        sheet_name = create_sheet_name(keywords, location, work_type)
        
        if group_key not in new_job_groups:
            new_job_groups[group_key] = {
                'jobs': [],
                'sheet_name': sheet_name,
                'job_ids': set()  # To track job IDs in this group
            }
        
        new_job_groups[group_key]['jobs'].append(job)
        new_job_groups[group_key]['job_ids'].add(job['job_id'])
    
    # Now load existing jobs and organize by sheet
    existing_sheets = {}
    
    if os.path.exists(excel_path):
        try:
            # Read Excel file - we need to read all sheets except 'Summary'
            xl = pd.ExcelFile(excel_path)
            sheet_names = [s for s in xl.sheet_names if s != 'Summary']
            
            # Read each sheet
            for sheet_name in sheet_names:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                
                # Convert DataFrame to list of dictionaries
                sheet_jobs = df.to_dict('records')
                
                # Create set of job IDs for this sheet
                job_ids = set()
                for job in sheet_jobs:
                    if 'job_id' in job and job['job_id'] and not pd.isna(job['job_id']):
                        job_ids.add(str(job['job_id']))
                
                existing_sheets[sheet_name] = {
                    'jobs': sheet_jobs,
                    'job_ids': job_ids
                }
                
            print(f"Loaded {len(sheet_names)} existing sheets from CRM")
            
        except Exception as e:
            print(f"Error loading existing jobs from Excel: {str(e)}")
            existing_sheets = {}
    
    # Merge new jobs with existing jobs, checking duplicates per sheet
    final_sheets = {}
    total_added = 0
    
    # First, copy all existing sheets to final
    for sheet_name, sheet_data in existing_sheets.items():
        final_sheets[sheet_name] = {
            'jobs': sheet_data['jobs'].copy(),
            'job_ids': sheet_data['job_ids'].copy()
        }
    
    # Now add new job groups, checking for duplicates within each sheet
    for group_key, group_data in new_job_groups.items():
        sheet_name = group_data['sheet_name']
        new_jobs_in_group = group_data['jobs']
        
        # If sheet doesn't exist yet in final, create it
        if sheet_name not in final_sheets:
            final_sheets[sheet_name] = {
                'jobs': [],
                'job_ids': set()
            }
        
        # Add non-duplicate jobs to this sheet
        added_to_sheet = 0
        for job in new_jobs_in_group:
            if job['job_id'] not in final_sheets[sheet_name]['job_ids']:
                # Add status fields for CRM functionality
                job['status'] = 'New'
                job['notes'] = ''
                job['date_added'] = datetime.datetime.now().strftime("%Y-%m-%d")
                
                final_sheets[sheet_name]['jobs'].append(job)
                final_sheets[sheet_name]['job_ids'].add(job['job_id'])
                added_to_sheet += 1
                total_added += 1
        
        if added_to_sheet > 0:
            print(f"Added {added_to_sheet} new jobs to sheet: {sheet_name}")
    
    if total_added == 0:
        print("No new jobs to add to CRM")
        return
    
    # Define fields for the Excel file
    fieldnames = [
        'job_id', 'job_title', 'company_name', 'location', 'publishing_date',
        'posted_time_ago', 'seniority_level', 'employment_type', 'job_function',
        'industries', 'status', 'notes', 'date_added', 'job_link', 'work_type_name'
    ]
    
    # Write to Excel
    try:
        # Create a styled Excel writer
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Create a summary sheet
            summary_data = []
            
            # Process each group and create a sheet
            for sheet_name, sheet_data in final_sheets.items():
                jobs = sheet_data['jobs']
                
                # Sort jobs by publishing date (newest first)
                sorted_jobs = sorted(jobs, 
                                     key=lambda x: x.get('publishing_date', '1970-01-01'), 
                                     reverse=True)
                
                # Add to summary data
                if sorted_jobs:
                    # Extract search parameters from the first job
                    first_job = sorted_jobs[0]
                    keywords = first_job.get('search_keywords', 'Unknown')
                    location = first_job.get('search_location', 'Unknown')
                    work_type_name = first_job.get('work_type_name', 'Any')
                    
                    search_name = f"{keywords.replace('%2B', '+')} in {location} ({work_type_name})"
                    
                    summary_data.append({
                        'Search': search_name,
                        'Jobs Found': len(sorted_jobs),
                        'Newest Job': sorted_jobs[0]['publishing_date'] if sorted_jobs else 'N/A',
                        'New Jobs': len([j for j in sorted_jobs if j['status'] == 'New']),
                        'Applied': len([j for j in sorted_jobs if j['status'] == 'Applied']),
                        'Interviews': len([j for j in sorted_jobs if j['status'] == 'Interview']),
                        'Offers': len([j for j in sorted_jobs if j['status'] == 'Offer']),
                        'Rejected': len([j for j in sorted_jobs if j['status'] == 'Rejected'])
                    })
                
                # Create a DataFrame for this group
                rows = []
                for job in sorted_jobs:
                    # Only include fields that are in fieldnames
                    row = {k: job.get(k, '') for k in fieldnames}
                    rows.append(row)
                
                df = pd.DataFrame(rows)
                
                # Write the sheet
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                # Get the worksheet
                worksheet = writer.sheets[sheet_name]
                
                # Format headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column width
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = max(max_length + 2, 10)
                    worksheet.column_dimensions[column].width = min(adjusted_width, 50)
                
                # Add filters to headers
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Color coding for status
                for row_idx, row in enumerate(df.iterrows(), 2):  # Start from row 2 (after header)
                    status = row[1].get('status', '')
                    status_cell = worksheet.cell(row=row_idx, column=fieldnames.index('status') + 1)
                    
                    if status == 'New':
                        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif status == 'Applied':
                        status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif status == 'Interview':
                        status_cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
                    elif status == 'Rejected':
                        status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif status == 'Offer':
                        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            # Create summary sheet
            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, index=False, sheet_name='Summary')
                
                # Format summary sheet
                summary_sheet = writer.sheets['Summary']
                
                # Format headers
                for col_num, column in enumerate(summary_df.columns, 1):
                    cell = summary_sheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column width
                for col in summary_sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = max(max_length + 2, 10)
                    summary_sheet.column_dimensions[column].width = min(adjusted_width, 50)
                
                # Add filters to headers
                summary_sheet.auto_filter.ref = summary_sheet.dimensions
        
        print(f"Total jobs added: {total_added}")
        print(f"Total jobs in CRM: {sum(len(sheet['jobs']) for sheet in final_sheets.values())}")
        print(f"CRM data saved to {excel_path} with {len(final_sheets)} search groups")
        
    except Exception as e:
        print(f"Error saving to Excel file: {str(e)}")
        # Fallback to CSV if Excel save fails
        csv_path = excel_path.replace('.xlsx', '.csv')
        
        # Create a single DataFrame with all jobs
        all_rows = []
        for sheet_data in final_sheets.values():
            for job in sheet_data['jobs']:
                row = {k: job.get(k, '') for k in fieldnames}
                all_rows.append(row)
        
        pd.DataFrame(all_rows).to_csv(csv_path, index=False)
        print(f"Saved to CSV instead at {csv_path}")

def create_sheet_name(keywords, location, work_type):
    """Create a simplified sheet name from search parameters"""
    # Get work type name
    work_type_name = config.WORK_TYPE_NAMES.get(work_type, "Any")
    
    # Replace %2B with + in keywords
    keywords = keywords.replace('%2B', '+')
    
    # Create base name
    if len(keywords) <= 10 and len(location) <= 10:
        # If short enough, use the full name
        sheet_name = f"{keywords}-{location}-{work_type_name}"
    else:
        # Truncate long names
        sheet_name = f"{keywords[:10]}-{location[:10]}-{work_type_name}"
    
    # Remove special characters not allowed in Excel sheet names
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '')
    
    # Ensure sheet name is no longer than 31 characters (Excel limit)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    
    return sheet_name

def save_jobs_to_file(jobs, filename=config.JSON_OUTPUT_PATH):
    """Saves job data to a JSON file"""
    # Ensure output directory exists
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    import json
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(jobs, f, ensure_ascii=False, indent=2)
    print(f"Job data saved to {filename}") 